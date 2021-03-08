
#   Openpyxl

#T# Table of contents

#T# Basic usage
#T# Workbooks
#T# Worksheets
#T# Cells

#T# Beginning of content

# |-------------------------------------------------------------
#T# Openpyxl is a package to read, and write spreadsheet files

#T# files can be edited in a semi interactive way, to see changes made to an open spreadsheet file, execute the openpyxl statements with the save function being last, and reload the spreadsheet file to see the changes
# |-------------------------------------------------------------

#T# Basic usage

# |-------------------------------------------------------------
#T# import the whole package, or import specific modules, functions, classes
from openpyxl import Workbook

#T# create a workbook
workbook1 = Workbook()

#T# get or set the active worksheet, the same as selecting it
worksheet1 = workbook1.active
workbook1.active = 0

#T# create new worksheets
worksheet2 = workbook1.create_sheet("sheet2_created")

#T# save the workbook in the file system hierarchy
workbook1.save("/tmp/spreadsheet1.xlsx")

# |-------------------------------------------------------------

#T# Workbooks

# |-------------------------------------------------------------
#T# set a workbook to be a template
workbook1.template = True
workbook1.save("/tmp/spreadsheet1.xltx")

#T# load a file from disc with the load_workbook function
from openpyxl import load_workbook
workbook1 = load_workbook("/tmp/spreadsheet1.xlsx")
# |-------------------------------------------------------------

#T# Worksheets

# |-------------------------------------------------------------

#T# a workbook is a dictionary of worksheets, where the keys are the names of the sheets and the values are the sheets themselves
worksheet2 = workbook1["sheet2_created"]

#T# get or set the title of a worksheet
str1 = worksheet2.title # 'sheet2_created'

#T# get or set the tab color of a worksheet, using a 6 hex RGB number
worksheet2.sheet_properties.tabColor = "C234FB"

#T# the sheetnames attribute of a workbook contains a list with the sheet names
list1 = workbook1.sheetnames

#T# a workbook is an iterable, its elements are the worksheets
for sheet_i1 in workbook1:
    pass

#T# a worksheet's content can be copied (not the charts, etc.)
worksheet1 = workbook1.copy_worksheet(worksheet2) # worksheet1.title == 'sheet2_created Copy'
# |-------------------------------------------------------------

#T# Cells

# |-------------------------------------------------------------
#T# a worksheet is a dictionary of cells, where the keys are the names of the cells and the values are the cell values
worksheet1['C2'] = "value in C2"
cellC2 = worksheet1['C2'] # <Cell 'sheet2_created Copy'.C2>

#T# access a cell using row and column numbers
cellD10 = worksheet1.cell(10, 4, "value in D10") # <Cell 'sheet2_created Copy'.D10>

#T# the value attribute of a cell is used to get or set the value in the cell
str1 = cellC2.value # 'value in C2'

#T# a range of cells is assigned with the slice operator, it's a tuple of cells
cells_B3_C4 = worksheet1['B3':'C4'] # the tuple has the cells (('B3', 'C3'), ('B4', 'C4'))

#T# whole rows and columns can be assigned as well
cells_D_E = worksheet1['D':'E'] # the actual range is up to the biggest row used, e.g. if it is row 11 then this is the same as the range 'D1':'E11'

#T# the iter_rows, iter_cols functions return a range, ordered by rows, columns respectively
cells_C5_E6 = worksheet1.iter_rows(min_row = 5, max_row = 6, min_col = 3, max_col = 5) # generator object Worksheet._cells_by_row, the cells are (('C5', 'D5', 'E5'), ('C6', 'D6', 'E6'))
cells_C5_E6 = worksheet1.iter_cols(min_row = 5, max_row = 6, min_col = 3, max_col = 5) # generator object Worksheet._cells_by_col, the cells are (('C5', 'C6'), ('D5', 'D6'), ('E5', 'E6'))

worksheet1 = workbook1.create_sheet(); worksheet1['C2'] = 5
#T# the rows, columns attributes have all the cells in the sheet, ordered by rows, columns respectively
cells_by_row1 = worksheet1.rows # (('A1', 'B1', 'C1'), ('A2', 'B2', 'C2'))
cells_by_col1 = worksheet1.columns # (('A1', 'A2'), ('B1', 'B2'), ('C1', 'C2'))

#T# the values property has all the cell values in the sheet, ordered by row
values_by_row1 = worksheet1.values # ((None, None, None), (None, None, 5))
# |-------------------------------------------------------------